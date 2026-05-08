from __future__ import annotations

from openpyxl import Workbook

from reportertool.metadata import extract_workbook_metadata
from reportertool.normalize import normalize_workbook
from reportertool.questionnaire_mapping import QuestionnaireItem, QuestionnaireMapping
from reportertool.xlsx_reader import XlsxWorkbook


def test_split_question_workbook_builds_base_answer_fact_and_question_table(tmp_path) -> None:
    path = tmp_path / "split.xlsx"
    wb = Workbook()
    front_page = wb.active
    front_page.title = "首页"
    front_page.append(["项目名称", "课程实施监测"])
    question = wb.create_sheet("1")
    question.append(
        [
            "问题id",
            "字段id",
            "字段名称",
            "字段取值",
            "维度id",
            "维度名称",
            "学科维度",
            "年级维度",
            "学期维度",
            "学校维度",
            "区域维度",
            "用户维度",
            "状态",
        ]
    )
    question.append(["Q1", "F1", "field_1", "1", "D1", "dim_1", "历史", "七年级", "2024下", "第一中学", "一区", "U001", "有效"])
    field_mapping = wb.create_sheet("字段映射关系")
    field_mapping.append(["题号", "题id", "字段id", "字段名称", "字段标题"])
    field_mapping.append([1, "Q1", "F1", "field_1", "同意"])
    dimension_mapping = wb.create_sheet("维度映射关系")
    dimension_mapping.append(["题号", "题id", "维度id", "维度名称", "维度组名称", "维度标题"])
    dimension_mapping.append([1, "Q1", "D1", "dim_1", "态度", "态度维度"])
    wb.save(path)
    workbook = XlsxWorkbook(path)

    result = normalize_workbook(
        workbook,
        title_metadata={
            "project_name": "课程实施监测",
            "form_title": "历史教师课程实施与教材使用情况表",
            "form_id": "FORM1",
            "role": "教师",
            "subject": "历史",
            "semester": "2024下",
        },
        questionnaire_mapping=mapping_for_history(),
        workbook_metadata=extract_workbook_metadata(workbook),
        source_file=path,
    )

    fact = result["base_answer_fact"][0]
    assert set(fact) >= {
        "fact_id",
        "dataset_id",
        "role",
        "subject",
        "user_id",
        "school",
        "region",
        "grade",
        "semester",
        "question_no",
        "local_question_key",
        "norm_id",
        "question_type",
        "question_text",
        "canonical_question",
        "answer_component_id",
        "answer_component_type",
        "answer_component_label",
        "field_value",
        "raw_value",
        "field_id",
        "field_name",
        "field_title",
        "dimension_id",
        "dimension_name",
        "source_file",
        "source_sheet",
        "source_row_index",
    }
    assert fact["role"] == "教师"
    assert fact["subject"] == "历史"
    assert fact["user_id"] == "U001"
    assert fact["school"] == "第一中学"
    assert fact["region"] == "一区"
    assert fact["grade"] == "七年级"
    assert fact["question_no"] == "1"
    assert fact["local_question_key"] == "教师|历史|1"
    assert fact["norm_id"] == "QG0001"
    assert fact["answer_component_id"] == "Q1:F1:D1"
    assert fact["answer_component_type"] == "option"
    assert fact["answer_component_label"] == "同意"
    assert fact["field_value"] == "1"
    assert fact["raw_value"] == "1"
    assert fact["source_sheet"] == "1"
    assert fact["source_row_index"] == "2"

    assert result["question_table"] == [
        {
            "entity_type": "question",
            "dataset_id": fact["dataset_id"],
            "role": "教师",
            "subject": "历史",
            "question_no": "1",
            "local_question_key": "教师|历史|1",
            "norm_id": "QG0001",
            "question_type": "单选题",
            "question_text": "是否使用新教材",
            "canonical_question": "是否使用新教材",
            "display_title": "是否使用新教材",
            "answer_component_id": "",
            "answer_component_type": "",
            "answer_component_label": "",
            "field_id": "",
            "field_name": "",
            "field_title": "",
            "dimension_id": "",
            "dimension_name": "",
            "dimension_group_name": "",
            "dimension_title": "",
            "join_key": "QG0001",
            "option_order": "",
        },
        {
            "entity_type": "answer_component",
            "dataset_id": fact["dataset_id"],
            "role": "教师",
            "subject": "历史",
            "question_no": "1",
            "local_question_key": "教师|历史|1",
            "norm_id": "QG0001",
            "question_type": "单选题",
            "question_text": "是否使用新教材",
            "canonical_question": "是否使用新教材",
            "display_title": "同意",
            "answer_component_id": "Q1:F1:D1",
            "answer_component_type": "option",
            "answer_component_label": "同意",
            "field_id": "F1",
            "field_name": "field_1",
            "field_title": "同意",
            "dimension_id": "D1",
            "dimension_name": "dim_1",
            "dimension_group_name": "态度",
            "dimension_title": "态度维度",
            "join_key": "QG0001|Q1:F1:D1",
            "option_order": "1",
        },
    ]
    assert result["quality_checks"] == []


def test_split_question_workbook_keeps_answer_columns_when_mappings_are_missing(tmp_path) -> None:
    path = tmp_path / "without_mapping.xlsx"
    wb = Workbook()
    wb.active.title = "首页"
    question = wb.create_sheet("2")
    question.append(["字段取值", "用户维度", "学校维度"])
    question.append(["课堂记录.docx", "U002", "第二中学"])
    wb.save(path)
    workbook = XlsxWorkbook(path)

    result = normalize_workbook(
        workbook,
        title_metadata={"role": "教师", "subject": "历史", "semester": "2024下"},
        questionnaire_mapping=mapping_for_history(),
        workbook_metadata=extract_workbook_metadata(workbook),
        source_file=path,
    )

    fact = result["base_answer_fact"][0]
    assert fact["question_no"] == "2"
    assert fact["norm_id"] == "QG0002"
    assert fact["answer_component_id"] == ""
    assert fact["answer_component_type"] == ""
    assert fact["answer_component_label"] == ""
    assert fact["field_id"] == ""
    assert fact["dimension_id"] == ""
    assert fact["raw_value"] == "课堂记录.docx"
    assert result["question_table"][0]["entity_type"] == "question"
    assert len(result["question_table"]) == 1


def test_unmatched_questions_emit_quality_check_and_use_local_question_key(tmp_path) -> None:
    path = tmp_path / "unmatched.xlsx"
    wb = Workbook()
    wb.active.title = "首页"
    question = wb.create_sheet("99")
    question.append(["字段取值", "用户维度"])
    question.append(["未知答案", "U099"])
    wb.save(path)
    workbook = XlsxWorkbook(path)

    result = normalize_workbook(
        workbook,
        title_metadata={"role": "教师", "subject": "历史"},
        questionnaire_mapping=mapping_for_history(),
        workbook_metadata=extract_workbook_metadata(workbook),
        source_file=path,
    )

    fact = result["base_answer_fact"][0]
    assert fact["norm_id"] == "教师|历史|99"
    assert result["quality_checks"] == [
        {
            "check_type": "unmatched_question",
            "severity": "warning",
            "source_sheet": "99",
            "source_row_index": "2",
            "local_question_key": "教师|历史|99",
            "message": "Question not found in normalized questionnaire mapping.",
        }
    ]


def test_wide_table_builds_facts_from_question_number_columns(tmp_path) -> None:
    path = tmp_path / "wide.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append(["用户维度", "学校维度", "区域维度", "1"])
    sheet.append(["U003", "第三中学", "三区", "是"])
    wb.save(path)
    workbook = XlsxWorkbook(path)

    result = normalize_workbook(
        workbook,
        title_metadata={"role": "教师", "subject": "历史"},
        questionnaire_mapping=mapping_for_history(),
        workbook_metadata=extract_workbook_metadata(workbook),
        source_file=path,
    )

    fact = result["base_answer_fact"][0]
    assert fact["question_no"] == "1"
    assert fact["user_id"] == "U003"
    assert fact["school"] == "第三中学"
    assert fact["region"] == "三区"
    assert fact["raw_value"] == "是"
    assert fact["source_sheet"] == "Sheet1"


def mapping_for_history() -> QuestionnaireMapping:
    return QuestionnaireMapping(
        [
            QuestionnaireItem(
                norm_id="QG0001",
                subject="历史",
                q_no="1",
                role="教师",
                question_type="单选题",
                question_text="是否使用新教材",
                canonical_question="是否使用新教材",
            ),
            QuestionnaireItem(
                norm_id="QG0002",
                subject="历史",
                q_no="2",
                role="教师",
                question_type="上传题",
                question_text="上传课堂记录",
                canonical_question="上传课堂记录",
            ),
        ]
    )
