from __future__ import annotations

from openpyxl import Workbook

from reportertool.metadata import extract_workbook_metadata
from reportertool.xlsx_reader import XlsxWorkbook


def test_extracts_front_page_and_question_type_stats(tmp_path) -> None:
    path = tmp_path / "metadata.xlsx"
    wb = Workbook()
    front_page = wb.active
    front_page.title = "首页"
    front_page.append(["项目名称", "课程实施监测"])
    front_page.append(["表单标题", "历史教师课程实施与教材使用情况表"])
    front_page.append(["填写数量", 12])
    front_page.append(["区域维度", "内蒙古自治区"])
    front_page.append(["学校维度", "第一中学"])
    stats = wb.create_sheet("题型统计")
    stats.append(["题号", "题型"])
    stats.append([1, "单选题"])
    stats.append([2, "上传题"])
    wb.save(path)

    metadata = extract_workbook_metadata(XlsxWorkbook(path))

    assert metadata["front_page"] == {
        "项目名称": "课程实施监测",
        "表单标题": "历史教师课程实施与教材使用情况表",
        "填写数量": "12",
        "区域维度": "内蒙古自治区",
        "学校维度": "第一中学",
    }
    assert metadata["scope"] == {
        "区域维度": "内蒙古自治区",
        "学校维度": "第一中学",
        "学科维度": "",
        "年级维度": "",
        "学期维度": "",
        "用户维度": "",
    }
    assert metadata["question_types"] == {"1": "单选题", "2": "上传题"}


def test_extracts_answer_components_from_field_and_dimension_mappings(tmp_path) -> None:
    path = tmp_path / "components.xlsx"
    wb = Workbook()
    wb.active.title = "首页"
    field_mapping = wb.create_sheet("字段映射关系")
    field_mapping.append(["题号", "题id", "字段id", "字段名称", "字段标题"])
    field_mapping.append([1, "Q1", "F1", "field_1", "同意"])
    dimension_mapping = wb.create_sheet("维度映射关系")
    dimension_mapping.append(["题号", "题id", "维度id", "维度名称", "维度组名称", "维度标题"])
    dimension_mapping.append([1, "Q1", "D1", "dim_1", "态度", "态度维度"])
    wb.save(path)

    metadata = extract_workbook_metadata(XlsxWorkbook(path))

    assert metadata["answer_components"] == [
        {
            "question_no": "1",
            "question_id": "Q1",
            "answer_component_id": "Q1:F1:D1",
            "answer_component_type": "option",
            "answer_component_label": "同意",
            "field_id": "F1",
            "field_name": "field_1",
            "field_title": "同意",
            "dimension_id": "D1",
            "dimension_name": "dim_1",
            "dimension_group_name": "态度",
            "dimension_title": "态度维度",
        }
    ]


def test_answer_component_columns_are_kept_when_mapping_sources_are_missing(tmp_path) -> None:
    path = tmp_path / "partial_components.xlsx"
    wb = Workbook()
    wb.active.title = "首页"
    field_mapping = wb.create_sheet("字段映射关系")
    field_mapping.append(["题号", "题id", "字段id", "字段名称", "字段标题"])
    field_mapping.append([3, "Q3", "F3", "upload_file", "课堂学习评价材料"])
    wb.save(path)

    metadata = extract_workbook_metadata(XlsxWorkbook(path))

    assert metadata["answer_components"] == [
        {
            "question_no": "3",
            "question_id": "Q3",
            "answer_component_id": "Q3:F3",
            "answer_component_type": "upload",
            "answer_component_label": "课堂学习评价材料",
            "field_id": "F3",
            "field_name": "upload_file",
            "field_title": "课堂学习评价材料",
            "dimension_id": "",
            "dimension_name": "",
            "dimension_group_name": "",
            "dimension_title": "",
        }
    ]
