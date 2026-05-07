from __future__ import annotations

import zipfile

from openpyxl import Workbook

from reportertool.xlsx_reader import XlsxWorkbook


def test_xlsx_workbook_reads_sheet_names_and_rows(tmp_path) -> None:
    path = tmp_path / "workbook.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "首页"
    ws.append(["项目名称", "课程实施监测"])
    ws.append(["填写数量", 3])
    for name in ("1", "字段映射关系", "维度映射关系", "题型统计"):
        wb.create_sheet(name)
    wb.save(path)

    workbook = XlsxWorkbook(path)

    assert workbook.sheet_names() == ["首页", "1", "字段映射关系", "维度映射关系", "题型统计"]
    assert workbook.rows("首页")[:2] == [["项目名称", "课程实施监测"], ["填写数量", "3"]]
    assert workbook.rows("缺失") == []


def test_xlsx_workbook_reads_rows_when_exported_dimension_is_a1(tmp_path) -> None:
    path = tmp_path / "bad_dimension.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "首页"
    ws.append(["项目名称", "课程实施监测"])
    ws.append(["表单标题", "历史教师课程实施与教材使用情况表"])
    wb.save(path)
    wb.close()

    rewritten_path = rewrite_sheet_dimension(path, "A1:A1")

    workbook = XlsxWorkbook(rewritten_path)

    assert workbook.rows("首页")[:2] == [
        ["项目名称", "课程实施监测"],
        ["表单标题", "历史教师课程实施与教材使用情况表"],
    ]


def rewrite_sheet_dimension(path, dimension: str):
    rewritten = path.with_suffix(".rewritten.xlsx")
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(rewritten, "w") as dst:
        for item in src.infolist():
            content = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = content.decode("utf-8")
                text = text.replace('ref="A1:B2"', f'ref="{dimension}"')
                content = text.encode("utf-8")
            dst.writestr(item, content)
    return rewritten
