from __future__ import annotations

from openpyxl import Workbook

from reportertool.input_detect import detect_workbook_type
from reportertool.xlsx_reader import XlsxWorkbook


def save_workbook(path, sheet_names: list[str], first_row: list[str] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_names[0]
    if first_row:
        ws.append(first_row)
    for name in sheet_names[1:]:
        wb.create_sheet(name)
    wb.save(path)


def test_detects_split_question_workbook_without_mapping_sheets(tmp_path) -> None:
    path = tmp_path / "split.xlsx"
    save_workbook(path, ["首页", "1", "题型统计"])

    result = detect_workbook_type(XlsxWorkbook(path))

    assert result == {
        "workbook_type": "split_question_workbook",
        "has_front_page": "true",
        "has_question_type_stats": "true",
        "numeric_sheet_count": "1",
    }


def test_detects_wide_table_when_first_row_has_multiple_fields(tmp_path) -> None:
    path = tmp_path / "wide.xlsx"
    save_workbook(path, ["Sheet1"], ["学校", "学科", "题1"])

    result = detect_workbook_type(XlsxWorkbook(path))

    assert result["workbook_type"] == "wide_table"
    assert result["first_row_field_count"] == "3"
