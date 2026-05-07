from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


class XlsxWorkbook:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        self._workbook = load_workbook(self.path, read_only=True, data_only=True)

    def sheet_names(self) -> list[str]:
        return list(self._workbook.sheetnames)

    def rows(self, sheet_name: str) -> list[list[str]]:
        if sheet_name not in self._workbook.sheetnames:
            return []
        sheet = self._workbook[sheet_name]
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([cell_to_text(value) for value in row])
        return rows


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
