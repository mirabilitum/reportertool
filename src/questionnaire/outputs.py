from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportertool.review_package import write_stage_status
from .components import build_option_review_rows, has_required_fill_marker
from .constants import WRITE_AUDIT_REVIEW_TABLES_ENV
from .manual_review import build_manual_candidate_group_rows
from .models import Question
from .text_normalize import env_flag

def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

def safe_cell(value: object) -> str | int | float:
    return "" if value is None else value  # type: ignore[return-value]

def autosize(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells[:200]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

def write_outputs(
    questions: list[Question],
    groups: list[list[int]],
    review_candidates: list[dict],
    auto_option_merge_rows: list[dict],
    manual_normalize_candidates: list[dict],
    manual_merge_rows: list[dict],
    output_dir: Path,
    input_paths: list[Path],
) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    review_dir = output_dir / "review"
    review_dir.mkdir(parents=True, exist_ok=True)
    write_audit_review_tables = env_flag(WRITE_AUDIT_REVIEW_TABLES_ENV)
    generated_review_csvs = [
        "answer_option_differences.csv",
        "auto_normalized_by_options.csv",
        "manual_normalize_applied.csv",
        "manual_normalize_candidates.csv",
        "question_dependencies.csv",
        "questionnaire_anomalies.csv",
    ]
    if not write_audit_review_tables:
        for name in generated_review_csvs:
            try:
                (review_dir / name).unlink(missing_ok=True)
            except OSError:
                pass
    subjects = sorted({q.subject for q in questions})

    extracted = {
        "source_count": len({q.source_file for q in questions}),
        "question_count": len(questions),
        "component_count": sum(len(q.answer_components) for q in questions),
        "subjects": subjects,
        "questions": [q.as_json() for q in questions],
    }
    questions_json = output_dir / "questions_extracted.json"
    questions_json.write_text(json.dumps(extracted, ensure_ascii=False, indent=2), encoding="utf-8")

    wide_rows: list[dict] = []
    long_rows: list[dict] = []
    component_rows: list[dict] = []
    option_set_rows: list[dict] = []
    option_difference_rows: list[dict] = []
    dependency_rows: list[dict] = []
    manual_candidate_group_rows = build_manual_candidate_group_rows(manual_normalize_candidates)
    for group_no, group in enumerate(groups, start=1):
        group_questions = [questions[i] for i in group]
        subject_count = len({q.subject for q in group_questions})
        question_text_counts = Counter(q.question_text for q in group_questions)
        canonical = max(
            group_questions,
            key=lambda q: (question_text_counts[q.question_text], len(q.question_text), -q.q_no),
        )
        group_match_method = "normalized_exact" if len({q.normalized_key for q in group_questions}) == 1 else "auto_option_or_fill_similarity"
        norm_id = f"QG{group_no:04d}"
        by_subject: dict[str, list[Question]] = defaultdict(list)
        for q in group_questions:
            by_subject[q.subject].append(q)
        row = {
            "norm_id": norm_id,
            "subject_count": subject_count,
            "role": canonical.role,
            "question_type": canonical.question_type,
            "canonical_question": canonical.question_text,
            "normalized_text": canonical.normalized_text,
            "match_method": group_match_method,
        }
        for subject in subjects:
            qs = sorted(by_subject.get(subject, []), key=lambda q: q.q_no)
            row[subject] = "; ".join(str(q.q_no) for q in qs)
        wide_rows.append(row)

        for q in sorted(group_questions, key=lambda q: (q.role, q.subject, q.q_no)):
            long_rows.append(
                {
                    "norm_id": norm_id,
                    "subject_count": subject_count,
                    "role": q.role,
                    "subject": q.subject,
                    "q_no": q.q_no,
                    "question_type": q.question_type,
                    "question_required": q.question_required,
                    "dependency_text": " | ".join(str(d["dependency_text"]) for d in q.dependencies),
                    "depends_on_q_no": " | ".join(str(d["depends_on_q_no"]) for d in q.dependencies),
                    "depends_on_option_orders": " | ".join(str(d["depends_on_option_orders"]) for d in q.dependencies),
                    "question_text": q.question_text,
                    "canonical_question": canonical.question_text,
                    "normalized_text": q.normalized_text,
                    "source_file": q.source_file,
                    "source_path": q.source_path,
                    "match_method": group_match_method,
                    "answer_component_count": len(q.answer_components),
                }
            )
            for component in q.answer_components:
                data = component.as_json()
                data["norm_id"] = norm_id
                data["role"] = q.role
                data["question_type"] = q.question_type
                data["question_required"] = q.question_required
                data["text_entry_required"] = has_required_fill_marker(component)
                data["question_text"] = q.question_text
                data["canonical_question"] = canonical.question_text
                component_rows.append(data)
            for dependency in q.dependencies:
                dependency_rows.append(
                    {
                        "norm_id": norm_id,
                        "role": q.role,
                        "subject": q.subject,
                        "q_no": q.q_no,
                        "question_type": q.question_type,
                        "question_required": q.question_required,
                        "question_text": q.question_text,
                        "canonical_question": canonical.question_text,
                        "dependency_text": dependency["dependency_text"],
                        "dependency_logic": dependency["dependency_logic"],
                        "depends_on_q_no": dependency["depends_on_q_no"],
                        "depends_on_option_orders": dependency["depends_on_option_orders"],
                        "dependency_required": dependency["dependency_required"],
                        "source_file": q.source_file,
                    }
                )
        group_option_set_rows, group_option_difference_rows = build_option_review_rows(
            norm_id,
            subject_count,
            group_questions,
            canonical,
        )
        option_set_rows.extend(group_option_set_rows)
        option_difference_rows.extend(group_option_difference_rows)

    wide_fields = [
        "norm_id",
        "subject_count",
        "role",
        "question_type",
        "canonical_question",
        "normalized_text",
        "match_method",
        *subjects,
    ]
    long_fields = [
        "norm_id",
        "subject_count",
        "role",
        "subject",
        "q_no",
        "question_type",
        "question_required",
        "dependency_text",
        "depends_on_q_no",
        "depends_on_option_orders",
        "question_text",
        "canonical_question",
        "normalized_text",
        "source_file",
        "source_path",
        "match_method",
        "answer_component_count",
    ]
    component_fields = [
        "norm_id",
        "role",
        "subject",
        "q_no",
        "question_type",
        "question_required",
        "question_text",
        "canonical_question",
        "component_id",
        "component_type",
        "component_label",
        "component_value",
        "option_order",
        "text_entry_required",
        "row_label",
        "col_label",
        "source_kind",
        "source_file",
    ]
    option_set_fields = [
        "norm_id",
        "subject_count",
        "role",
        "subject",
        "q_no",
        "question_type",
        "question_text",
        "canonical_question",
        "question_required",
        "dependency_text",
        "depends_on_q_no",
        "depends_on_option_orders",
        "option_relation",
        "option_variant_count",
        "minimum_reference_subject",
        "minimum_reference_q_no",
        "component_count",
        "option_count",
        "matrix_row_count",
        "matrix_col_count",
        "upload_count",
        "scalar_count",
        "minimum_reference_components",
        "full_components",
        "extra_vs_minimum",
        "missing_vs_minimum",
        "options",
        "matrix_rows",
        "matrix_cols",
        "upload_components",
        "scalar_components",
        "text_entry_options",
        "required_text_entry_options",
        "normalized_component_key",
        "source_file",
    ]
    option_difference_fields = [
        "norm_id",
        "role",
        "subject",
        "q_no",
        "question_type",
        "canonical_question",
        "question_required",
        "dependency_text",
        "depends_on_q_no",
        "depends_on_option_orders",
        "option_relation",
        "option_variant_count",
        "minimum_reference_subject",
        "minimum_reference_q_no",
        "minimum_reference_components",
        "current_components",
        "extra_vs_minimum",
        "missing_vs_minimum",
        "minimum_reference_text_entry_options",
        "current_text_entry_options",
        "source_file",
    ]
    dependency_fields = [
        "norm_id",
        "role",
        "subject",
        "q_no",
        "question_type",
        "question_required",
        "question_text",
        "canonical_question",
        "dependency_text",
        "dependency_logic",
        "depends_on_q_no",
        "depends_on_option_orders",
        "dependency_required",
        "source_file",
    ]
    candidate_fields = [
        "similarity",
        "role",
        "type",
        "subject_a",
        "q_no_a",
        "text_a",
        "subject_b",
        "q_no_b",
        "text_b",
    ]
    option_normalize_review_fields = [
        "decision_status",
        "human_decision",
        "review_note",
        "similarity",
        "topic_similarity",
        "topic_jaccard",
        "semantic_provider",
        "semantic_similarity",
        "llm_should_merge",
        "llm_confidence",
        "llm_reason",
        "llm_risk_flags",
        "reason",
        "role",
        "question_type",
        "option_relation",
        "source_group_a",
        "source_group_b",
        "source_file_a",
        "subject_a",
        "q_no_a",
        "text_a",
        "components_a",
        "fill_blank_count_a",
        "source_file_b",
        "subject_b",
        "q_no_b",
        "text_b",
        "components_b",
        "fill_blank_count_b",
        "missing_from_b_vs_a",
        "extra_in_b_vs_a",
    ]
    manual_candidate_group_fields = [
        "标注说明",
        "merge_id",
        "candidate_group_id",
        "suggested_action",
        "review_logic",
        "risk_flags",
        "role",
        "question_type",
        "pair_count",
        "question_count",
        "max_similarity",
        "max_topic_similarity",
        "max_topic_jaccard",
        "semantic_provider_summary",
        "max_semantic_similarity",
        "llm_decision_summary",
        "llm_reason_summary",
        "llm_risk_flags_summary",
        "option_relation_summary",
        "reason_summary",
        "subjects_qnos",
        "question_keys_json",
        "question_texts",
        "components_by_question",
        "difference_summary",
    ]
    manual_merge_fields = [
        "source_row",
        "manual_label",
        "applied_label",
        "status",
        "source_file_a",
        "subject_a",
        "q_no_a",
        "text_a",
        "source_file_b",
        "subject_b",
        "q_no_b",
        "text_b",
        "merge_id",
        "human_decision",
        "review_note",
    ]

    wide_path = output_dir / "normalized_question_mapping_wide.csv"
    long_path = output_dir / "normalized_question_mapping_long.csv"
    components_path = output_dir / "normalized_answer_components_long.csv"
    option_sets_path = output_dir / "normalized_answer_option_sets.csv"
    candidates_path = output_dir / "near_match_review_candidates.csv"
    anomalies_path = review_dir / "questionnaire_anomalies.csv"
    option_differences_path = review_dir / "answer_option_differences.csv"
    dependencies_path = review_dir / "question_dependencies.csv"
    auto_option_merge_path = review_dir / "auto_normalized_by_options.csv"
    manual_normalize_candidates_path = review_dir / "manual_normalize_candidates.csv"
    manual_normalize_candidate_groups_path = review_dir / "manual_normalize_candidate_groups.csv"
    manual_merge_validation_path = review_dir / "manual_normalize_applied.csv"

    write_csv(wide_path, wide_rows, wide_fields)
    write_csv(long_path, long_rows, long_fields)
    write_csv(components_path, component_rows, component_fields)
    write_csv(option_sets_path, option_set_rows, option_set_fields)
    write_csv(manual_normalize_candidate_groups_path, manual_candidate_group_rows, manual_candidate_group_fields)
    if write_audit_review_tables:
        write_csv(candidates_path, review_candidates, candidate_fields)
        write_csv(option_differences_path, option_difference_rows, option_difference_fields)
        write_csv(dependencies_path, dependency_rows, dependency_fields)
        write_csv(auto_option_merge_path, auto_option_merge_rows, option_normalize_review_fields)
        write_csv(manual_normalize_candidates_path, manual_normalize_candidates, option_normalize_review_fields)
        write_csv(manual_merge_validation_path, manual_merge_rows, manual_merge_fields)
    else:
        try:
            candidates_path.unlink(missing_ok=True)
        except OSError:
            pass

    anomalies = []
    for q in questions:
        if q.question_type in ("单选题", "多选题", "量表题", "表格组合题") and not q.answer_components:
            anomalies.append(
                {
                    "severity": "warning",
                    "source_file": q.source_file,
                    "role": q.role,
                    "subject": q.subject,
                    "q_no": q.q_no,
                    "question_type": q.question_type,
                    "issue": "no_answer_components",
                    "question_text": q.question_text,
                }
            )
    anomaly_fields = ["severity", "source_file", "role", "subject", "q_no", "question_type", "issue", "question_text"]
    if write_audit_review_tables:
        write_csv(anomalies_path, anomalies, anomaly_fields)

    workbook_path = output_dir / "normalized_question_mapping.xlsx"
    wb = Workbook()
    sheets = [
        ("mapping_wide", wide_fields, wide_rows),
        ("mapping_long", long_fields, long_rows),
        ("answer_components", component_fields, component_rows),
        ("option_sets", option_set_fields, option_set_rows),
        ("manual_candidate_groups", manual_candidate_group_fields, manual_candidate_group_rows),
    ]
    if anomalies:
        sheets.append(("anomalies", anomaly_fields, anomalies))
    if write_audit_review_tables:
        sheets.extend(
            [
                ("option_differences", option_difference_fields, option_difference_rows),
                ("question_dependencies", dependency_fields, dependency_rows),
                ("auto_normalized", option_normalize_review_fields, auto_option_merge_rows),
                ("manual_candidates", option_normalize_review_fields, manual_normalize_candidates),
                ("manual_applied", manual_merge_fields, manual_merge_rows),
                ("review_candidates", candidate_fields, review_candidates),
            ]
        )
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for idx, (name, fields, rows) in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet(name)
        ws.title = name
        ws.append(fields)
        for row in rows:
            ws.append([safe_cell(row.get(f)) for f in fields])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        autosize(ws)

    ws_summary = wb.create_sheet("summary")
    summary_rows = [
        ("source_files", len({q.source_file for q in questions})),
        ("subjects", len(subjects)),
        ("questions", len(questions)),
        ("answer_components", len(component_rows)),
        ("answer_option_sets", len(option_set_rows)),
        ("answer_option_differences", len(option_difference_rows)),
        ("question_dependencies", len(dependency_rows)),
        ("auto_option_normalized_pairs", len(auto_option_merge_rows)),
        ("manual_normalize_candidates", len(manual_normalize_candidates)),
        ("manual_normalize_candidate_groups", len(manual_candidate_group_rows)),
        ("manual_normalize_applied_rows", len(manual_merge_rows)),
        ("normalized_groups", len(groups)),
        ("groups_in_2plus_subjects", sum(1 for r in wide_rows if r["subject_count"] >= 2)),
        ("near_match_candidates", len(review_candidates)),
        ("anomalies", len(anomalies)),
    ]
    for item in summary_rows:
        ws_summary.append(item)
    autosize(ws_summary)
    wb.save(workbook_path)

    output_paths = [
        wide_path,
        long_path,
        components_path,
        option_sets_path,
        manual_normalize_candidate_groups_path,
        workbook_path,
        questions_json,
    ]
    if write_audit_review_tables:
        output_paths.extend(
            [
                option_differences_path,
                dependencies_path,
                auto_option_merge_path,
                manual_normalize_candidates_path,
                manual_merge_validation_path,
                candidates_path,
                anomalies_path,
            ]
        )

    status = "passed_with_warnings" if anomalies else "passed"
    status_path = write_stage_status(
        output_dir,
        stage_name="normalize-questionnaires",
        status=status,
        warning_count=len(anomalies),
        input_paths=input_paths,
        output_paths=output_paths,
        quality_checks_path=anomalies_path if write_audit_review_tables else workbook_path,
        next_human_action="Review manual_normalize_candidate_groups.csv and fill merge_id; details are in normalized_question_mapping.xlsx.",
        extra={
            "source_files": len(input_paths),
            "questions": len(questions),
            "answer_components": len(component_rows),
            "answer_option_sets": len(option_set_rows),
            "answer_option_differences": len(option_difference_rows),
            "question_dependencies": len(dependency_rows),
            "auto_option_normalized_pairs": len(auto_option_merge_rows),
            "manual_normalize_candidates": len(manual_normalize_candidates),
            "manual_normalize_candidate_groups": len(manual_candidate_group_rows),
            "manual_normalize_applied_rows": len(manual_merge_rows),
            "normalized_groups": len(groups),
            "near_match_candidates": len(review_candidates),
        },
    )

    return {
        "status": status,
        "source_files": len(input_paths),
        "questions": len(questions),
        "answer_components": len(component_rows),
        "answer_option_sets": len(option_set_rows),
        "answer_option_differences": len(option_difference_rows),
        "question_dependencies": len(dependency_rows),
        "auto_option_normalized_pairs": len(auto_option_merge_rows),
        "manual_normalize_candidates": len(manual_normalize_candidates),
        "manual_normalize_candidate_groups": len(manual_candidate_group_rows),
        "manual_normalize_applied_rows": len(manual_merge_rows),
        "normalized_groups": len(groups),
        "near_match_candidates": len(review_candidates),
        "warning_count": len(anomalies),
        "stage_status": str(status_path),
    }
